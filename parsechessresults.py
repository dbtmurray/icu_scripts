import urllib.request
import bs4
import sys
from enum import Enum
import io
from openpyxl import load_workbook

"""
script for parsing chess-results.com pages into ICU-CSV format.

Usage:

python3 parsechessresults.py "http://chess-results.com/tnr367947.aspx?lan=1&art=9&fedb=IRL&fed=IRL&turdet=YES&flag=30&snr=42" # an individual
python3 parsechessresults.py "http://chess-results.com/tnr373918.aspx?lan=1&art=20&fed=IRL&flag=30" # a team
python3 parsechessresults.py "http://chess-results.com/tnr385901.aspx?lan=1&zeilen=0&art=25&fedb=IRL&turdet=YES&flag=30&prt=4&excel=2010" # a team, from Excel file
python3 parsechessresults.py "http://www.4nclresults.co.uk/2018-19/4ncl/1/2b/export/" 12 # parse 4NCL site rounds 1 and 2 for div 2b
python3 parsechessresults.py "http://www.4nclresults.co.uk/2018-19/4ncl/7/2b/export/" 72b,82c #4ncl that spans multiple divisions, here 7/2b and 8/2c

Limitations:
    -need to add ICU code for the player (output as ???? instead)
    -may split surnames from first names wrongly. Maybe we could get them from the FIDE site instead
"""

class Colour(Enum):
    UNKNOWN = 0
    WHITE = 1
    BLACK = 2


class PlayerResult:
    def __init__(self, player, rd, score=0, colour=Colour.UNKNOWN, opp_name="", opp_rating=0, opp_title=None, opp_fed=None):
        self.player = player
        self.rd = rd
        self.score = score
        self.colour = colour
        self.opp_name = opp_name
        self.opp_rating = opp_rating
        if is_fide_title(opp_title):
            self.opp_title = opp_title
        else:
            self.opp_title = ""
        self.opp_fed = opp_fed
        
class Player:
    def __init__(self, name, icu_code="????"):
        self.name = name
        self.icu_code = icu_code
        self.score = 0
        self.results = {}


def is_opponent_class(css_class):
    if css_class is None:
        return False
    if (str(css_class).startswith(("CRg", "CRng"))) and " " in str(css_class):
        return True
    elif str(css_class) in ["CRg2", "CRg1"]:
        return True
    else:
        return False

def is_header_class(css_class):
    """return True iff we think this is the right class for the column headers"""
    return str(css_class) in ["CRg1b", "CRng1b"]


def commaize(name):
    tokens = name.split()
    surnames = 1
    if len(tokens) > 3 and " ".join(tokens[:2]).lower() in ["van der", "van den", "mac an"]:
        surnames = 3
    elif len(tokens) > 2 and tokens[0].lower() in ["mc", "mac", "al", "de", "ui", "ni", "o", "ul", "vam", "van", "der", "den"]:
        surnames = 2
    elif len(tokens) > 2 and " ".join(tokens[:2]) in ["Plaza Reino"]:
        surnames = 2
    return " ".join(tokens[:surnames]) + "," + " ".join(tokens[surnames:])

def is_fide_title(text):
    titles = ["GM", "IM", "FM", "CM", "WGM", "WIM", "WFM", "WCM"]
    return text in titles

def parse_4ncl_title(text):
    text = text.replace("j", "").strip()
    text = text.replace("*", "").strip()
    titles_dict = {"" : "", "w" : "", "c" : "CM", "f" : "FM", "i" : "IM", "g" : "GM", "wc" : "WCM", "wf" : "WFM", "wi" : "wim", "wg" : "WGM"}
    return titles_dict[text]


# these 3 methods take a ChessResults score that looks like "w 1"

def score_character(result):
    if result[-1] == "0":
        return "0"
    elif result[-1] == "1":
        return "1"
    else:
        return "="

def score_value(result):
    if result[-1] == "K":
        result = result[:-1]
    if result[-1] == "0":
        return 0
    elif result[-1] == "1":
        return 1
    else:
        return 0.5

def score_colour(result):
    if result[0] in "wW":
        return Colour.WHITE
    elif result[0] in "bBsS":
        return Colour.BLACK
    else:
        return None

# these two methods take a result from the 4NCL format, which looks like "1 - 0"
def score_character_4ncl(result):
    if result[0] == "0":
        return "0"
    elif result[0] == "1":
        return "1"
    else:
        return "="

def score_value_4ncl(result):
    if result[0] == "0":
        return 0
    elif result[0] == "1":
        return 1
    else:
        return 0.5


def colour_character(colour):
    if colour == Colour.WHITE:
        return "W"
    elif colour == Colour.BLACK:
        return "B"
    else:
        return "-"

def reverse_colour(colour):
    if colour == Colour.WHITE:
        return Colour.BLACK
    elif colour == Colour.BLACK:
        return Colour.WHITE
    else:
        raise ValueError("could not reverse colour for %s" % colour)

def merge_players(playerlist1, playerlist2):
    """merges the players destructively into playerlist1"""
    for player2 in playerlist2:
        match = [player1 for player1 in playerlist1 if player1.name == player2.name]
        if match:
            player1 = match[0]
            player1.results.update(player2.results)
            player1.score += player2.score
        else:
            playerlist1.append(player2)
    return playerlist1


def parse_4ncl(soup, rd):
    TEAM_NAME = "Gonzaga" # pass in on command line?
    players = []
    td = soup.find_all("td",text=TEAM_NAME)[1]
    # check if we are looking at the first named team or the second one
    team_line = td.parent.find_all("td")
    if td == team_line[1]:
        first_team = True
    elif td == team_line[4]:
        first_team = False
    else:
        raise ValueError("could not identify if we are looking at the home or away team")
    tr = td.parent.next_sibling.next_sibling
    while len(tr.find_all("td")) == 9:
        tds = tr.find_all("td")
        if first_team:
            name = tds[2].text.replace(", ", ",")
            opp_name = tds[6].text
            opp_title = parse_4ncl_title(tds[7].text)
            opp_rating = int(tds[8].text.split()[0])
            colour = score_colour(tds[1].text)
            result = tds[5].text.strip()
        else:
            name = tds[6].text.replace(", ", ",")
            opp_name = tds[2].text
            opp_title = parse_4ncl_title(tds[3].text)
            opp_rating = int(tds[4].text.split()[0])
            colour = reverse_colour(score_colour(tds[1].text))
            # reverse the result to get it from our point of view
            result = tds[5].text.strip()[::-1]
        player = Player(name)
        player.score += score_value_4ncl(result)
        score = score_character_4ncl(result)
        players.append(player)
        opp_fed = "ENG" # would be nice to be able to look these up, but it's maybe not necessary
        player_result = PlayerResult(player, rd, score, colour, opp_name, opp_rating, opp_title, opp_fed)
        player.results[rd] = player_result
        tr = tr.next_sibling.next_sibling
    return players


def parse_team(soup):
    players = []
    playerInfo = soup.find(text="Player info") or soup.find(text=lambda t:t.startswith("Player details for"))
    tr = playerInfo.find_next("tr",class_="CRg1b")
    output_lines = []
    player = None
    while tr is not None:
        tds = list(tr.children)
        if tds and "Rd." in tds[0].text:
            pass
        else:
            if len(tds) == 1:
                tokens = tds[0].text.split()
                nameTokens = []
                #if isFideTitle(nameTokens[0]):
                    #nameTokens = nameTokens[1:]
                for token in tokens:
                    if token.isnumeric() or is_fide_title(token):
                        break
                    else:
                        nameTokens.append(token)
                name = " ".join(nameTokens)
                name = commaize(name)
                player = Player(name)
                players.append(player)
            elif len(tds) == 10:
                # Rd, SNo, title, name, rating, fed, Rp, Pts, result, board
                rd = int(tds[0].text)
                title = tds[2].text
                name = commaize(tds[3].text)
                rating = tds[4].text
                if int(rating) == 0:
                    rating = ""
                fed = tds[5].text
                result = tds[8].text
                score = score_character(result)
                player.score += score_value(result)
                colour = score_colour(result)
                playerResult = PlayerResult(player, rd, score, colour, name, rating, title, fed)
                player.results[rd] = playerResult
            elif len(tds) == 9: # no Rp field!
                # Rd, SNo, title, name, rating, fed, Rp, Pts, result, board
                rd = int(tds[0].text)
                title = tds[2].text
                name = commaize(tds[3].text)
                rating = tds[4].text
                if int(rating) == 0:
                    rating = ""
                fed = tds[5].text
                result = tds[7].text
                score = score_character(result)
                player.score += score_value(result)
                colour = score_colour(result)


                playerResult = PlayerResult(player, rd, score, colour, name, rating, title, fed)
                player.results[rd] = playerResult

        tr = tr.find_next_sibling("tr")
    return players



def parse_individual(soup):

    div = [div for div in soup.find_all("div",class_="defaultDialog") if div.find("h2").text == "Player info"][0]

    player = div.find(text="Name").next.text.strip()
    playerName = commaize(player)
    player = Player(playerName)
    output_lines = []

    trs = div.find_all("tr", class_=is_opponent_class)

    total = 0
    output_lines.append("Player,????,%s" % playerName)

    for tr in trs:
        tds = [el for el in list(tr.children) if isinstance(el, bs4.element.Tag)]
        rd = int(tds[0].text)
        title = tds[3].text
        if title in ["AIM", "AFM", "AGM", "ACM"]: # not recognized by ICU software
            title = ""
        name = commaize(tds[4].text)
        rating = tds[5].text
        if int(rating) == 0:
            rating = ""
        fed = tds[6].text
        result = tds[8].text
        score = score_character(result)
        total += score_value(result)
        colour = score_colour(result)

        output_lines.append("%d,%s,%s,%s,%s,%s,%s" % (rd, score, colour, name, rating, title, fed))
        player.score += score_value(result)
        playerResult = PlayerResult(player, rd, score, colour, name, rating, title, fed)
        player.results[rd] = playerResult
    return player

def parse_individual_auto(soup):
    """Parse results for an individual player, automatically getting the right columns
    from the headers.
    Should replace parse_individual and parse_team with this approach"""

    div = [div for div in soup.find_all("div",class_="defaultDialog") if div.find("h2").text == "Player info"][0]

    player = div.find(text="Name").next.text.strip()
    playerName = commaize(player)
    player = Player(playerName)
    output_lines = []

    header = div.find_all("tr", class_=is_header_class)[0]
    cols = [el.text for el in list(header.children) if isinstance(el, bs4.element.Tag)]
    round_index = cols.index("Rd.")
    opp_name_index = cols.index("Name")
    try:
        rating_index = cols.index("RtgI")
    except ValueError:
        rating_index = cols.index("Rtg")
    fed_index = cols.index("FED")
    result_index = cols.index("Res.")
    title_index = cols.index("") # maybe this won't always get it

    trs = div.find_all("tr", class_=is_opponent_class)

    total = 0
    output_lines.append("Player,????,%s" % playerName)

    for tr in trs:
        tds = [el for el in list(tr.children) if isinstance(el, bs4.element.Tag)]
        rd = int(tds[round_index].text)
        title = tds[title_index].text
        if title in ["AIM", "AFM", "AGM", "ACM"]: # not recognized by ICU software
            title = ""
        name = commaize(tds[opp_name_index].text)
        rating = tds[rating_index].text
        if int(rating) == 0:
            rating = ""
        fed = tds[fed_index].text
        result = tds[result_index].text
        score = score_character(result)
        total += score_value(result)
        colour = score_colour(result)

        output_lines.append("%d,%s,%s,%s,%s,%s,%s" % (rd, score, colour, name, rating, title, fed))
        player.score += score_value(result)
        playerResult = PlayerResult(player, rd, score, colour, name, rating, title, fed)
        player.results[rd] = playerResult
    return player



def parse_team_from_xlsx(workbook):
    players = []
    if "PlayerInfo" in workbook.sheetnames:
        ws = workbook["PlayerInfo"]
        row = 2
        while ws["A%s" % row].value:
            cell_value = ws["A%s" % row].value
            if isinstance(cell_value, str):
                if "Rd." not in cell_value:
                    tokens = cell_value.split()
                    name_tokens = []
                    for token in tokens:
                        if token.isnumeric():
                            break
                        if is_fide_title(token):
                            break
                        name_tokens.append(token)
                    name = " ".join(name_tokens).strip()
                    player = Player(name)
                    players.append(player)
            elif isinstance(cell_value, int):
                rd = cell_value
                opp_title = ws["C%s" % row].value or ""
                opp_name = ws["D%s" % row].value.strip()
                opp_rating = ws["E%s" % row].value
                opp_fed = ws["F%s" % row].value
                result = ws["I%s" % row].value.strip()
                score = score_character(result)
                player.score += score_value(result)
                colour = score_colour(result)

                if result[-1] != "K": # walkover or other unplayed game
                    player_result = PlayerResult(player, rd, score, colour, opp_name, opp_rating, opp_title, opp_fed)
                    player.results[rd] = player_result
            row += 1
        return players
    elif "PlayerDetails" in workbook.sheetnames:
        ws = workbook["PlayerDetails"]
        row = 3
        while ws["A%s" % row].value:
            cell_value = ws["A%s" % row].value
            if isinstance(cell_value, str):
                if "Rd." not in cell_value:
                    tokens = cell_value.split()
                    name_tokens = []
                    for token in tokens:
                        if token.isnumeric():
                            break
                        if is_fide_title(token):
                            continue
                        name_tokens.append(token)
                    name = " ".join(name_tokens).strip()
                    player = Player(name)
                    players.append(player)
            elif isinstance(cell_value, int):
                rd = cell_value
                opp_title = ws["C%s" % row].value or ""
                opp_name = ws["D%s" % row].value.strip()
                opp_rating = ws["E%s" % row].value
                opp_fed = ws["F%s" % row].value
                result = ws["H%s" % row].value.strip()
                score = score_character(result)
                player.score += score_value(result)
                colour = score_colour(result)

                player_result = PlayerResult(player, rd, score, colour, opp_name, opp_rating, opp_title, opp_fed)
                player.results[rd] = player_result
            row += 1
        return players

def parse_commas_from_player_pairings(workbook):
    """This method figures out where commas go
    in player names. So it's maybe optional.
    Returns a dict of {name without comma : name with comma}"""
    if "PlayerPairings" not in workbook.sheetnames:
        return None
    ws = workbook["PlayerPairings"]
    commas = {}
    row = 4
    # sometimes has ratings in this format, in which case a different column is used
    if isinstance(ws["D4"].value, int):
        name2_col = "G"
    else:
        name2_col = "F"

    while ws["A%s" % row].value:
        name1 = ws["C%s" % row].value
        name2 = ws["%s%s" % (name2_col, row)].value
        if name1 and "," in name1:
            # strip trailing (w) or (b)
            if name1.endswith(("(w)", "(b)")):
                name1 = name1[:-4]
            name1_no_comma = name1.replace(",", "")
            commas[name1_no_comma] = name1
        if name2 and "," in name2:
            if name2.endswith(("(w)", "(b)")):
                name2 = name2[:-4]
            name2_no_comma = name2.replace(",", "")
            commas[name2_no_comma] = name2
        
        row += 1

    return commas

def apply_commas(players, commas):
    if commas is None:
        for player in players:
            player.name = commaize(player.name)
            for rd, result in player.results.items():
                result.opp_name = commaize(result.opp_name)
    else:
        for player in players:
            player.name = commas[player.name]
            for rd, result in player.results.items():
                result.opp_name = commas[result.opp_name]


def parse(source, rounds=None):
    url = source

    if not url.startswith("http"):
        # local file. Only for CR individual result right now
        with open(url, "r") as f:
            data = f.read()
            soup = bs4.BeautifulSoup(data, 'html.parser')

        event = soup.title.text.split(" - ")[1].strip()

        if "Team composition" in str(data) or "Player overview for" in str(data):
            players = parse_team(soup)
        else:
            players = [parse_individual_auto(soup)]

    if "chess-results.com" in url and "excel=" in url:
        response = urllib.request.urlopen(url)
        xlsx = response.read()
        wb = load_workbook(io.BytesIO(xlsx))
        #wb = load_workbook("womens.xlsx")

        players = parse_team_from_xlsx(wb)
        commas = parse_commas_from_player_pairings(wb)
        apply_commas(players, commas)
        
        if "TeamComposition" in wb.sheetnames:
            event = wb["TeamComposition"]["A2"].value
        else:
            event = wb["Sheet1"]["A2"].value

    elif "chess-results.com" in url:
        response = urllib.request.urlopen(url)
        data = response.read()
        soup = bs4.BeautifulSoup(data, 'html.parser')
        event = soup.title.text.split(" - ")[1].strip()

        if "Team composition" in str(data) or "Player overview for" in str(data):
            #print("Parsing team")
            players = parse_team(soup)
        else:
            #print("Parsing for individual")
            players = [parse_individual_auto(soup)]

    elif "4nclresults.co.uk" in url:
        if rounds is None:
            raise ValueError("Need to specify round numbers for 4NCL results")
        players = []
        urlparts = url.split("/")
        index = urlparts.index("4ncl") + 1
        if len(rounds) > 3:
            # like "72b,82c" to get rounds 7/2b and 8/2c
            new_rds = rounds.split(",")
            new_rds = [[rd[:-2], rd[-2:]] for rd in new_rds]
        else:
            new_rds = [[rd] for rd in rounds]

        for rd in new_rds:
            new_urlparts = urlparts
            new_urlparts[index:index+len(rd)] = rd
            url = "/".join(new_urlparts)
            response = urllib.request.urlopen(url)
            data = response.read()
            soup = bs4.BeautifulSoup(data, 'html.parser')
            round_players = parse_4ncl(soup, int(rd[0]))
            players = merge_players(players, round_players)
            event = "4NCL Rounds %s-%s" % (rounds[0], rounds[-1])

    if not players:
        raise ValueError("Could not parse any players")

    return event, players


def output(event, players, url):
    output_lines = []
    max_round = max(rd for player in players for rd in player.results)
    min_round = min(rd for player in players for rd in player.results)
    num_rounds = max_round - min_round + 1

    output_lines.append("Event,%s" % event)
    output_lines.append("Start,??/??/20??")
    output_lines.append("End,??/??/20??")
    output_lines.append("Rounds,%d" % num_rounds)
    output_lines.append("Website,%s" % url)

    for player in players:
        output_lines.append("")
        name = player.name.replace(", ", ",")
        output_lines.append("Player,%s,%s" % (player.icu_code, name))
        for rd in range(min_round, max_round + 1):
            result = player.results.get(rd)
            # adjust so the first round rated is always reported to the ICU as round 1
            # even if it was a later round in the tournament (typical for 4ncl)
            adjusted_rd = rd - min_round + 1
            if not result:
                output_lines.append("%d,0,-" % adjusted_rd)
            else:
                if result.opp_rating:
                    opp_rating = str(result.opp_rating)
                else:
                    opp_rating = ""
                opp_name = result.opp_name.replace(", ", ",")
                output_lines.append("%d,%s,%s,%s,%s,%s,%s" % (
                    adjusted_rd, result.score, colour_character(result.colour),
                    opp_name, opp_rating, result.opp_title, result.opp_fed))
        output_lines.append("Total,%3.1f" % player.score)
    print("\n".join(output_lines))

if __name__ == "__main__":

    source = sys.argv[1]
    rds = None
    if len(sys.argv) > 2:
        rds = sys.argv[2]
    event, players = parse(source, rds)
    output(event, players, source)

